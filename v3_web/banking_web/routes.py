import csv
import io
import json
import os
from datetime import datetime, date

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, Response
from functools import wraps

_USE_PG = os.environ.get("DATABASE_URL") is not None

routes_bp = Blueprint("routes", __name__)


# ── Helpers ─────────────────────────────────────────────────

def login_required(view):
    @wraps(view)
    def wrapped(**kwargs):
        if "user_id" not in session:
            flash("Please log in first", "warning")
            return redirect(url_for("auth.login"))
        return view(**kwargs)
    return wrapped


def _svc():
    from banking_core.services.atm_service import ATMService
    return ATMService()


def _usvc():
    from banking_core.services import UserService
    return UserService()


def _acct():
    return _svc().get_account(session["user_id"])


def _conn():
    if _USE_PG:
        from banking_core.data.postgres_adapter import get_pg_connection
        return get_pg_connection()
    from banking_core.data.db_manager import db
    return db.get_connection("ecosystem")


def _c():
    return _conn().cursor()


def _user():
    return _usvc().get_user(session["user_id"])


def _feature_context(user=None):
    """Real user features for ML models (mirrors CLI feature mapping)."""
    user = user or _user()
    c = _conn().cursor()
    c.execute("SELECT COUNT(*) FROM transactions WHERE user_id=? AND type IN ('withdraw','transfer')",
              (session["user_id"],))
    txn_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM fraud_flags WHERE user_id=?", (session["user_id"],))
    fraud_count = c.fetchone()[0]
    last_active = user.get("last_active") or ""
    days_inactive = 999
    if last_active:
        try:
            days_inactive = (date.today() - datetime.strptime(last_active[:10], "%Y-%m-%d").date()).days
        except Exception:
            days_inactive = 0
    return {
        "age": user.get("age") or 25,
        "age_group": user.get("age_group") or "Adult",
        "income_bracket": user.get("income_bracket") or "not_earning_student",
        "balance": user.get("balance") or 5000,
        "txn_count": txn_count,
        "fraud_count": fraud_count,
        "days_inactive": days_inactive,
        "bank": user.get("bank") or "",
        "credit_score": user.get("credit_score") or 600,
        "is_minor": bool(user.get("is_minor", False)),
    }


def _income_amount(bracket):
    """Map income bracket code to an approximate annual income figure (CLI-compatible)."""
    low_map = {
        "not_earning_student": 0, "not_earning_homemaker": 0, "not_earning_unemployed": 0,
        "not_earning_retired": 0, "earning_under_2.5L": 200000, "earning_2.5L_5L": 375000,
        "earning_5L_10L": 750000, "earning_10L_25L": 1750000, "earning_25L_plus": 3000000,
    }
    return low_map.get(bracket, 500000)


def _run_model(inst, predict_fn, train_fn=None):
    """Try predict; on failure try train (optional) then predict again. Returns (ok, result, error)."""
    try:
        result = predict_fn()
        if result is not None and (not isinstance(result, dict) or "error" not in result):
            return True, result, None
        error = result.get("error") if isinstance(result, dict) else "Empty result"
    except Exception as e:
        error = f"{type(e).__name__}: {e}"
    if train_fn is not None:
        try:
            train_fn()
            result = predict_fn()
            if result is not None and (not isinstance(result, dict) or "error" not in result):
                return True, result, None
            error = result.get("error") if isinstance(result, dict) else "Empty result"
        except Exception as e:
            error = f"{type(e).__name__}: {e}"
    return False, None, error


def _snapshot(name, bank=None, metric=None, kind="json"):
    """Read a precomputed ML/report snapshot from Neon (None when not in PG mode)."""
    if not _USE_PG:
        return None
    try:
        from banking_core.data.postgres_adapter import get_ml_snapshot
        return get_ml_snapshot(name, bank=bank, metric=metric, kind=kind)
    except Exception:
        return None


def _user_report():
    """Per-user precomputed report; computes + persists on miss (write-through)."""
    if not _USE_PG:
        return None
    key = f"report_user_{session['user_id']}"
    rep = _snapshot(key, kind="report")
    if rep:
        return rep
    try:
        from banking_core.analytics.report_generator import compute_user_report, store_user_report
        rep = compute_user_report(session["user_id"])
        if rep:
            try:
                store_user_report(session["user_id"], rep)
            except Exception:
                pass
        return rep
    except Exception:
        return None


def _train_report(train_name, fn):
    try:
        metrics = fn()
        if isinstance(metrics, dict) and metrics.get("skipped"):
            return train_name, "Skipped", str(metrics.get("skipped"))[:60]
        if isinstance(metrics, dict) and metrics.get("error"):
            return train_name, "Failed", metrics["error"][:80]
        if isinstance(metrics, dict):
            key = metrics.get(next(iter(metrics), ""), "") if metrics else ""
            return train_name, "Trained", f"{str(key)[:60]}"
        return train_name, "Trained", ""
    except Exception as e:
        return train_name, "Failed", str(e)[:80]


# ── Index ────────────────────────────────────────────────────

@routes_bp.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("routes.dashboard"))
    return redirect(url_for("auth.login"))


# ── Dashboard ────────────────────────────────────────────────

@routes_bp.route("/dashboard")
@login_required
def dashboard():
    account = _acct()
    if not account:
        flash("Account not found", "error")
        return redirect(url_for("auth.logout"))
    result = _svc().mini_statement(session["user_id"], limit=8)
    transactions = result.get("transactions", [])
    feats = _feature_context()
    c = _c()
    c.execute("SELECT COUNT(*) as n, COALESCE(SUM(amount),0) as tot FROM transactions WHERE user_id=? AND type IN ('deposit','credit')", (session["user_id"],))
    stats = dict(zip([d[0] for d in c.description], c.fetchone()))
    c.execute("SELECT COUNT(*) as n, COALESCE(SUM(amount),0) as tot FROM transactions WHERE user_id=? AND type IN ('withdraw','transfer')", (session["user_id"],))
    spend = dict(zip([d[0] for d in c.description], c.fetchone()))

    churn = None
    try:
        from banking_core.models import ChurnPredictor
        cp = ChurnPredictor()
        ok, res, _ = _run_model(cp, lambda: cp.predict({
            "age": feats["age"], "income_bracket": feats["income_bracket"],
            "balance": account.balance, "txn_count": feats["txn_count"],
            "days_inactive": feats["days_inactive"]}), cp.train)
        churn = res if ok else None
    except Exception:
        churn = None

    try:
        from banking_core.models.model_monitor import ModelMonitor
        from banking_core.auto_retrain import AutoRetrainScheduler
        monitor = ModelMonitor()
        stale = monitor.get_stale_models(max_days=7)
    except Exception:
        stale = []
    return render_template("dashboard.html", account=account, transactions=transactions,
                           feats=feats, stats=stats, spend=spend, churn=churn, stale=stale)


# ── ATM Operations ───────────────────────────────────────────

@routes_bp.route("/balance")
@login_required
def balance():
    account = _acct()
    if not account:
        return redirect(url_for("auth.logout"))
    result = _svc().get_balance(session["user_id"])
    mini = _svc().mini_statement(session["user_id"], limit=6)
    return render_template("balance.html", account=account, info=result,
                           transactions=mini.get("transactions", []))


@routes_bp.route("/deposit", methods=["GET", "POST"])
@login_required
def deposit():
    account = _acct()
    if request.method == "POST":
        try:
            amount = float(request.form.get("amount", "").strip())
        except (ValueError, TypeError):
            flash("Invalid amount", "error")
            return render_template("deposit.html", account=account)
        result = _svc().deposit(session["user_id"], amount, channel="web")
        if "error" in result:
            flash(result["error"], "error")
            return render_template("deposit.html", account=account)
        flash(f"Deposited ₹{amount:,.2f}. New balance: ₹{result['balance_after']:,.2f} (+2 credit score)", "success")
        return redirect(url_for("routes.balance"))
    return render_template("deposit.html", account=account)


@routes_bp.route("/withdraw", methods=["GET", "POST"])
@login_required
def withdraw():
    account = _acct()
    if request.method == "POST":
        confirmed = request.form.get("confirmed") == "yes"
        try:
            amount = float(request.form.get("amount", "").strip())
        except (ValueError, TypeError):
            flash("Invalid amount", "error")
            return render_template("withdraw.html", account=account)

        fraud = None
        if not confirmed:
            try:
                fraud = _svc().check_fraud(session["user_id"], amount)
            except Exception:
                fraud = {"is_suspicious": False}

        if fraud and fraud.get("is_suspicious") and not confirmed:
            return render_template("withdraw_confirm.html", account=account, amount=amount, fraud=fraud)

        result = _svc().withdraw(session["user_id"], amount, channel="web")
        if "error" in result:
            flash(result["error"], "error")
            return render_template("withdraw.html", account=account)
        denoms = result.get("denominations", {})
        flash(f"Withdrew ₹{amount:,.2f}. Remaining: ₹{result['balance_after']:,.2f} — " +
              " + ".join(f"₹{d}×{n}" for d, n in sorted(denoms.items(), reverse=True)), "success")
        return redirect(url_for("routes.balance"))
    return render_template("withdraw.html", account=account)


@routes_bp.route("/transfer", methods=["GET", "POST"])
@login_required
def transfer():
    account = _acct()
    if request.method == "POST":
        try:
            amount = float(request.form.get("amount", "").strip())
        except (ValueError, TypeError):
            flash("Invalid amount", "error")
            return render_template("transfer.html", account=account)
        target = request.form.get("target_account", "").strip()
        mode = request.form.get("mode", "bank_transfer").strip()
        if not target:
            flash("Target account is required", "error")
            return render_template("transfer.html", account=account)
        result = _svc().transfer(session["user_id"], amount, target, is_upi=(mode == "upi"), channel="web")
        if "error" in result:
            flash(result["error"], "error")
            return render_template("transfer.html", account=account)
        flash(f"Transferred ₹{amount:,.2f} to {target} via {mode.upper()}", "success")
        return redirect(url_for("routes.balance"))
    return render_template("transfer.html", account=account)


@routes_bp.route("/statement")
@login_required
def statement():
    result = _svc().mini_statement(session["user_id"], limit=100)
    txns = result.get("transactions", [])
    credits = sum(t["amount"] for t in txns if t["type"] in ("deposit", "credit"))
    debits = sum(t["amount"] for t in txns if t["type"] in ("withdraw", "transfer"))
    return render_template("statement.html", transactions=txns, credits=credits, debits=debits)


# ── Profile & PIN ────────────────────────────────────────────

@routes_bp.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = _user()
    if not user:
        return redirect(url_for("auth.logout"))
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        if name and email and phone:
            _usvc().update_user(session["user_id"], name=name, email=email, phone=phone)
            session["user_name"] = name
            flash("Profile updated", "success")
        else:
            flash("All fields required", "error")
        return redirect(url_for("routes.profile"))
    return render_template("profile.html", user=user)


@routes_bp.route("/change-pin", methods=["GET", "POST"])
@login_required
def change_pin():
    if request.method == "POST":
        current = request.form.get("current_pin", "").strip()
        new_pin = request.form.get("new_pin", "").strip()
        confirm = request.form.get("confirm_pin", "").strip()
        if not current or not new_pin or not confirm:
            flash("All fields required", "error")
        elif new_pin != confirm:
            flash("New PINs do not match", "error")
        elif len(new_pin) != 4 or not new_pin.isdigit():
            flash("PIN must be exactly 4 digits", "error")
        else:
            result = _usvc().change_pin(session["user_id"], current, new_pin)
            if result is True:
                flash("PIN changed successfully", "success")
                return redirect(url_for("routes.profile"))
            flash(result, "error")
        return render_template("change_pin.html")
    return render_template("change_pin.html")


# ── Credit Score ─────────────────────────────────────────────

@routes_bp.route("/credit-score")
@login_required
def credit_score():
    account = _acct()
    if not account:
        return redirect(url_for("auth.logout"))
    c = _c()
    c.execute("SELECT * FROM credit_history WHERE user_id=? ORDER BY recorded_at DESC LIMIT 20",
              (session["user_id"],))
    cols = [d[0] for d in c.description]
    history = [dict(zip(cols, r)) for r in c.fetchall()]
    feats = _feature_context()
    breakdown = {
        "base": 600 if not account.is_minor else 650,
        "income": min(120, feats["income_bracket"].startswith("earning") * 40 + int(feats["income_bracket"].count("L") * 15)),
        "balance": min(80, int(account.balance / 50000) * 8),
        "activity": min(60, min(feats["txn_count"], 12) * 5),
        "fees": -min(40, feats["fraud_count"] * 10),
    }
    breakdown["total"] = max(300, min(900, breakdown["base"] + breakdown["income"] + breakdown["balance"] + breakdown["activity"] + breakdown["fees"]))
    rating = "Excellent" if breakdown["total"] >= 750 else "Good" if breakdown["total"] >= 650 else "Fair" if breakdown["total"] >= 550 else "Poor"
    return render_template("credit_score.html", account=account, history=history, breakdown=breakdown, rating=rating)


# ── Savings Goals ────────────────────────────────────────────

@routes_bp.route("/savings", methods=["GET", "POST"])
@login_required
def savings():
    conn = _conn()
    if request.method == "POST":
        name = request.form.get("goal_name", "").strip()
        target = request.form.get("target_amount", "").strip()
        deadline = request.form.get("deadline", "").strip()
        if name and target:
            try:
                conn.execute(
                    "INSERT INTO savings_goals (user_id, goal_name, target_amount, deadline) VALUES (?,?,?,?)",
                    (session["user_id"], name, float(target), deadline),
                )
                conn.commit()
                flash("Savings goal created!", "success")
            except Exception as e:
                flash(f"Error: {e}", "error")
        return redirect(url_for("routes.savings"))
    c = conn.cursor()
    c.execute("SELECT * FROM savings_goals WHERE user_id=? ORDER BY created_at DESC", (session["user_id"],))
    cols = [d[0] for d in c.description]
    return render_template("savings.html", goals=[dict(zip(cols, r)) for r in c.fetchall()])


@routes_bp.route("/savings/<int:goal_id>/add", methods=["POST"])
@login_required
def savings_add(goal_id):
    amount = request.form.get("amount", "").strip()
    if amount:
        _conn().execute("UPDATE savings_goals SET current_amount = current_amount + ? WHERE goal_id=? AND user_id=?",
                        (float(amount), goal_id, session["user_id"]))
        _conn().commit()
        flash("Amount added to goal!", "success")
    return redirect(url_for("routes.savings"))


@routes_bp.route("/savings/<int:goal_id>/delete", methods=["POST"])
@login_required
def savings_delete(goal_id):
    _conn().execute("DELETE FROM savings_goals WHERE goal_id=? AND user_id=?", (goal_id, session["user_id"]))
    _conn().commit()
    flash("Goal deleted", "info")
    return redirect(url_for("routes.savings"))


@routes_bp.route("/savings/optimize", methods=["GET", "POST"])
@login_required
def savings_optimize():
    conn = _conn()
    c = conn.cursor()
    c.execute("SELECT * FROM savings_goals WHERE user_id=? AND is_completed=FALSE", (session["user_id"],))
    cols = [d[0] for d in c.description]
    goals = [dict(zip(cols, r)) for r in c.fetchall()]
    account = _acct()
    result = None
    user = _user()

    if request.method == "POST":
        target = request.form.get("target_amount", "").strip()
        months = request.form.get("deadline_months", "12").strip()
        monthly_income = request.form.get("monthly_income", "").strip()
        monthly_expenses = request.form.get("monthly_expenses", "").strip()
        try:
            from banking_core.models import SavingsGoalOptimizer
            so = SavingsGoalOptimizer()
            result = so.optimize(
                target_amount=float(target) if target else 100000,
                deadline_months=int(months) if months else 12,
                current_balance=account.balance if account else 0,
                monthly_income=float(monthly_income) if monthly_income else _income_amount(user.get("income_bracket")) / 12,
                monthly_expenses=float(monthly_expenses) if monthly_expenses else 0,
            )
            if isinstance(result, dict) and result.get("error"):
                flash(result["error"], "error")
                result = None
        except Exception as e:
            flash(f"Optimizer error: {e}", "error")

    suggestions = []
    if goals and account:
        total_target = sum(g["target_amount"] for g in goals)
        total_current = sum(g["current_amount"] for g in goals)
        remaining = total_target - total_current
        for g in goals:
            progress = (g["current_amount"] / g["target_amount"] * 100) if g["target_amount"] > 0 else 0
            months_left = 12
            if g.get("deadline"):
                try:
                    dl = datetime.strptime(g["deadline"], "%Y-%m-%d").date()
                    months_left = max(1, round((dl - date.today()).days / 30.44))
                except Exception:
                    months_left = 12
            monthly = (g["target_amount"] - g["current_amount"]) / months_left if months_left else 0
            suggestions.append({**g, "progress": round(progress, 1), "suggested_monthly": round(monthly, 0), "months_left": months_left})
    return render_template("savings_optimize.html", goals=goals, suggestions=suggestions,
                           result=result, account=account)


# ── Loan Applications ────────────────────────────────────────

@routes_bp.route("/loans", methods=["GET", "POST"])
@login_required
def loans():
    conn = _conn()
    account = _acct()
    user = _user()

    if request.method == "POST":
        loan_type = request.form.get("loan_type", "personal")
        amount = request.form.get("amount", "").strip()
        tenure = request.form.get("tenure", "12").strip()
        if amount:
            try:
                from banking_core.models import LoanDefaultModel
                lm = LoanDefaultModel()
                ok, risk, err = _run_model(lm, lambda: lm.predict({
                    "credit_score": account.credit_score, "balance": account.balance,
                    "age": user["age"], "income_bracket": user.get("income_bracket")},
                    float(amount), 10.0, int(tenure)), lm.train)
                conn.execute(
                    """INSERT INTO loan_applications
                       (user_id, loan_type, amount_requested, tenure_months, status, risk_score, predicted_default)
                       VALUES (?,?,?,?,?,?,?)""",
                    (session["user_id"], loan_type, float(amount), int(tenure),
                     "approved" if ok and risk.get("risk_score", 1) < 0.5 else "pending",
                     risk.get("risk_score") if ok else None,
                     risk.get("risk_score") if ok else None),
                )
                conn.commit()
                flash("Loan application submitted!", "success")
            except Exception as e:
                flash(f"Error: {e}", "error")
        return redirect(url_for("routes.loans"))

    c = conn.cursor()
    c.execute("SELECT * FROM loan_applications WHERE user_id=? ORDER BY applied_at DESC", (session["user_id"],))
    cols = [d[0] for d in c.description]
    applications = [dict(zip(cols, r)) for r in c.fetchall()]

    loan_offers = []
    if account and not account.is_minor:
        from banking_core.models import LoanDefaultModel
        lm = LoanDefaultModel()
        base = {"credit_score": account.credit_score, "balance": account.balance,
                "age": user["age"], "income_bracket": user.get("income_bracket")}
        offers = [
            ("Personal Loan", "personal", 500000, 10.5, 24),
            ("Home Top-up", "home", 2000000, 8.5, 60),
            ("Overdraft", "overdraft", 250000, 12.0, 12),
            ("Education Loan", "education", 1000000, 9.5, 36),
        ]
        for label, key, max_amt, rate, tenure in offers:
            ok, risk, err = _run_model(lm, lambda: lm.predict(base, min(max_amt, 300000), rate, tenure), lm.train)
            risk_pct = round(risk.get("risk_score", 1) * 100, 1) if ok else None
            eligible = ok and risk_pct is not None and risk_pct < 60
            loan_offers.append({
                "type": label, "key": key, "rate": rate, "max": max_amt, "tenure": tenure,
                "risk_pct": risk_pct, "eligible": eligible, "error": err if not ok else None,
            })
    return render_template("loans.html", applications=applications, offers=loan_offers, account=account)


# ── Security / Fraud ─────────────────────────────────────────

@routes_bp.route("/security")
@login_required
def security():
    c = _c()
    c.execute("SELECT * FROM fraud_flags WHERE user_id=? ORDER BY flagged_at DESC LIMIT 20", (session["user_id"],))
    cols = [d[0] for d in c.description]
    flags = [dict(zip(cols, r)) for r in c.fetchall()]
    feats = _feature_context()
    return render_template("security.html", flags=flags, feats=feats)


# ── Reports & Export ─────────────────────────────────────────

@routes_bp.route("/reports")
@login_required
def reports():
    account = _acct()
    result = _svc().mini_statement(session["user_id"], limit=100)
    return render_template("reports.html", account=account, transactions=result.get("transactions", []))


@routes_bp.route("/export/csv")
@login_required
def export_csv():
    c = _c()
    c.execute("SELECT * FROM transactions WHERE user_id=? ORDER BY timestamp DESC", (session["user_id"],))
    cols = [d[0] for d in c.description]
    rows = c.fetchall()
    si = io.StringIO()
    w = csv.writer(si)
    w.writerow(cols)
    w.writerows(rows)
    return Response(si.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment;filename=transactions.csv"})


@routes_bp.route("/export/excel")
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    c = _c()
    c.execute("SELECT * FROM transactions WHERE user_id=? ORDER BY timestamp DESC", (session["user_id"],))
    cols = [d[0] for d in c.description]
    rows = c.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(cols)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        ws.append(list(row))
    for i, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(32, len(col) + 6))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows) + 1}"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return Response(bio.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=transactions.xlsx"})


@routes_bp.route("/export/passbook")
@login_required
def export_passbook():
    try:
        from banking_core.report_generator import ReportGenerator
        account = _acct()
        c = _c()
        c.execute("SELECT * FROM transactions WHERE user_id=? ORDER BY timestamp DESC", (session["user_id"],))
        cols = [d[0] for d in c.description]
        txns = [dict(zip(cols, r)) for r in c.fetchall()]
        user = _user()
        rg = ReportGenerator()
        user_data = {"user_id": account.user_id, "name": account.name, "email": account.email,
                     "phone": account.phone, "bank": account.bank, "account_no": account.account_number,
                     "card_no": account.card_number, "account_type": account.account_type,
                     "balance": account.balance, "credit_score": account.credit_score,
                     "age_group": account.age_group, "created_at": user.get("created_at", "")}
        pdf_path = rg.generate_passbook(user_data, txns=txns, prompt_open=False)
        if pdf_path and os.path.exists(str(pdf_path)):
            with open(str(pdf_path), "rb") as f:
                pdf_data = f.read()
            return Response(pdf_data, mimetype="application/pdf",
                            headers={"Content-Disposition": f"attachment;filename=passbook_{account.user_id}.pdf"})
        flash("Passbook generation failed", "error")
    except Exception as e:
        flash(f"Passbook error: {e}", "error")
    return redirect(url_for("routes.reports"))


@routes_bp.route("/export/summary-card")
@login_required
def export_summary_card():
    try:
        from banking_core.report_generator import ReportGenerator
        account = _acct()
        rg = ReportGenerator()
        card_path = rg.generate_account_summary_card(account)
        if card_path and os.path.exists(str(card_path)):
            with open(str(card_path), "rb") as f:
                data = f.read()
            return Response(data, mimetype="application/pdf",
                            headers={"Content-Disposition": f"attachment;filename=summary_card_{account.user_id}.pdf"})
        flash("Summary card generation failed", "error")
    except Exception as e:
        flash(f"Summary card error: {e}", "error")
    return redirect(url_for("routes.reports"))


# ── Investment Suggestions ───────────────────────────────────

@routes_bp.route("/investment-suggestions")
@login_required
def investment_suggestions():
    account = _acct()
    user = _user()
    if not account:
        return redirect(url_for("auth.logout"))
    risk = request.args.get("risk", "moderate").strip()
    if risk not in ("low", "moderate", "high"):
        risk = "moderate"
    suggestions = []
    total_yearly = 0
    error = None
    if risk == "moderate":
        rep = _user_report() or {}
        inv = rep.get("investments")
        if inv and inv.get("products"):
            suggestions = [{"category": p.get("name", p.get("category", "Product")), "products": [p],
                            "total_pct": p.get("allocation_pct", 0)} for p in inv["products"]]
            total_yearly = inv.get("total_yearly", 0)
            return render_template("investment.html", suggestions=suggestions, account=account,
                                   risk=risk, error=error, total_yearly=total_yearly)
    try:
        from banking_core.models.investment_recommender import InvestmentRecommender
        ir = InvestmentRecommender()
        result = ir.recommend(age=user["age"], income_bracket=user.get("income_bracket", "mid"),
                              balance=account.balance, risk_tolerance=risk)
        if isinstance(result, dict) and "products" in result:
            suggestions = [{"category": p.get("type", p.get("name", "Product")), "products": [p],
                            "total_pct": p.get("allocation_pct", p.get("allocation", p.get("pct", 0)))} for p in result["products"]]
            total_yearly = result.get("total_expected_return_yearly", 0)
        elif isinstance(result, list):
            suggestions = [{"category": p.get("type", p.get("name", "Product")), "products": [p],
                            "total_pct": p.get("allocation_pct", p.get("allocation", p.get("pct", 0)))} for p in result]
        else:
            error = str(result)
    except Exception as e:
        error = str(e)
    return render_template("investment.html", suggestions=suggestions, account=account,
                           risk=risk, error=error, total_yearly=total_yearly)


# ── RFM Segmentation ─────────────────────────────────────────

@routes_bp.route("/rfm")
@login_required
def rfm():
    account = _acct()
    c = _c()
    c.execute("SELECT * FROM transactions WHERE user_id=? ORDER BY timestamp DESC", (session["user_id"],))
    cols = [d[0] for d in c.description]
    txns = [dict(zip(cols, r)) for r in c.fetchall()]
    result = None
    error = None
    rep = _user_report() or {}
    result = rep.get("rfm")
    if not result:
        try:
            from banking_core.models import RFMSegmenter
            rfm = RFMSegmenter()
            if txns:
                result = rfm.segment(txns)
                behavior = rfm.get_segment_behavior(result.get("segment")) if result.get("segment") else None
                if behavior:
                    result["behavior"] = behavior
        except Exception as e:
            error = str(e)
    return render_template("rfm.html", account=account, result=result, error=error, txn_count=len(txns))


# ── Bank Explorer ────────────────────────────────────────────

@routes_bp.route("/bank-explorer")
@login_required
def bank_explorer():
    try:
        from banking_core.bank_attributes import BANK_PREFIXES, get_bank_attrs
        bank_names = sorted(BANK_PREFIXES.keys())
        banks = []
        for name in bank_names:
            attrs = get_bank_attrs(name)
            banks.append({"name": name, **attrs})
    except Exception:
        banks = []

    category = request.args.get("cat", "all").strip()
    q = request.args.get("q", "").strip().lower()
    if category != "all":
        banks = [b for b in banks if b.get("type") == category]
    if q:
        banks = [b for b in banks if q in b["name"].lower()]

    page = request.args.get("page", 1, type=int)
    per_page = 12
    total = len(banks)
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    paged = banks[start:start + per_page]

    random_bank = None
    if request.args.get("random"):
        import random
        random_bank = banks[random.randrange(len(banks))] if banks else None

    return render_template("bank_explorer.html", banks=paged, total=total, page=page,
                           pages=pages, category=category, q=q, random_bank=random_bank)


# ── Analytics ────────────────────────────────────────────────

def _da():
    from banking_core.data_analysis import DataAnalysis
    return DataAnalysis()


@routes_bp.route("/analytics")
@login_required
def analytics():
    account = _acct()
    c = _c()
    c.execute("SELECT COUNT(*) as txn_count, COALESCE(SUM(amount),0) as total_spent FROM transactions WHERE user_id=? AND type IN ('withdraw','transfer')",
              (session["user_id"],))
    spending = dict(zip([d[0] for d in c.description], c.fetchone()))
    c.execute("SELECT COUNT(*) as txn_count, COALESCE(SUM(amount),0) as total_dep FROM transactions WHERE user_id=? AND type IN ('deposit','credit')",
              (session["user_id"],))
    income = dict(zip([d[0] for d in c.description], c.fetchone()))
    c.execute("SELECT type, COUNT(*) as cnt, COALESCE(SUM(amount),0) as tot FROM transactions WHERE user_id=? GROUP BY type ORDER BY cnt DESC",
              (session["user_id"],))
    cols = [d[0] for d in c.description]
    breakdown = [dict(zip(cols, r)) for r in c.fetchall()]
    return render_template("analytics.html", account=account, spending=spending,
                           income=income, breakdown=breakdown)


@routes_bp.route("/analytics/monthly-trend")
@login_required
def monthly_trend():
    industry = None
    try:
        df = _da().monthly_trend()
        if df is not None and hasattr(df, "columns"):
            industry = df.to_dict(orient="records")
    except Exception:
        industry = None
    c = _c()
    date_sql = "TO_CHAR(timestamp, 'YYYY-MM')" if _USE_PG else "strftime('%Y-%m', timestamp)"
    c.execute(f"""SELECT {date_sql} as month, type, SUM(amount) as total, COUNT(*) as cnt
                 FROM transactions WHERE user_id=? GROUP BY month, type ORDER BY month""", (session["user_id"],))
    cols = [d[0] for d in c.description]
    user_trends = [dict(zip(cols, r)) for r in c.fetchall()]
    return render_template("monthly_trend.html", industry=industry, user_trends=user_trends)


@routes_bp.route("/analytics/channel-breakdown")
@login_required
def channel_breakdown():
    try:
        raw = _da().channel_breakdown()
        channels = [{"channel": k, "vol": v.get("Vol", 0), "val": v.get("Val", 0)} for k, v in raw.items()]
    except Exception:
        channels = None
    c = _c()
    c.execute("SELECT channel, COUNT(*) as cnt, COALESCE(SUM(amount),0) as tot FROM transactions WHERE user_id=? GROUP BY channel",
              (session["user_id"],))
    cols = [d[0] for d in c.description]
    user_channels = [dict(zip(cols, r)) for r in c.fetchall()]
    return render_template("channel_breakdown.html", channels=channels, user_channels=user_channels)


@routes_bp.route("/analytics/market-share")
@login_required
def market_share():
    ms = None
    tb = None
    try:
        da = _da()
        ms_df = da.market_share()
        tb_series = da.top_banks()
        if ms_df is not None and hasattr(ms_df, "columns"):
            ms = ms_df.to_dict(orient="records")
        if tb_series is not None and hasattr(tb_series, "items"):
            tb = [{"Bank_Name": k, "Total_Txn_Vol": v} for k, v in tb_series.items()]
    except Exception:
        pass
    return render_template("market_share.html", market_share=ms, top_banks=tb)


@routes_bp.route("/analytics/growth-rate")
@login_required
def growth_rate():
    gr = None
    try:
        da = _da()
        result = da.growth_rate()
        if result is not None and hasattr(result, "columns"):
            gr = result.to_dict(orient="records")
    except Exception:
        pass
    return render_template("growth_rate.html", growth_rate=gr)


@routes_bp.route("/analytics/user-vs-industry")
@login_required
def user_vs_industry():
    account = _acct()
    comparison = None
    rep = _user_report() or {}
    comparison = rep.get("user_vs_industry")
    if not comparison:
        try:
            user_data = {"bank": account.bank, "balance": account.balance, "age_group": account.age_group}
            comparison = _da().user_vs_industry(user_data)
        except Exception:
            pass
    return render_template("user_vs_industry.html", account=account, comparison=comparison)


@routes_bp.route("/analytics/bank-overview")
@login_required
def bank_overview():
    bank_name = request.args.get("bank", "").strip() or "STATE BANK OF INDIA"
    overview = None
    error = None
    try:
        da = _da()
        available = da.get_banks()
        if bank_name not in available:
            bank_name = available[0] if available else "STATE BANK OF INDIA"
        overview = da.bank_overview(bank_name)
    except Exception as e:
        error = str(e)
    return render_template("bank_overview.html", bank_name=bank_name, overview=overview,
                           error=error)


@routes_bp.route("/analytics/compare")
@login_required
def bank_compare():
    presets = {
        "PSU": ["STATE BANK OF INDIA", "BANK OF BARODA", "PUNJAB NATIONAL BANK", "CANARA BANK"],
        "Private": ["HDFC BANK LTD", "ICICI BANK LTD", "AXIS BANK LTD", "KOTAK MAHINDRA BANK LTD"],
        "Volume": ["STATE BANK OF INDIA", "HDFC BANK LTD", "ICICI BANK LTD", "AXIS BANK LTD"],
    }
    selected = []
    try:
        preset = request.args.get("preset", "")
        if preset in presets:
            selected = presets[preset]
        else:
            raw = request.args.get("banks", "").split(",")
            selected = [b.strip() for b in raw if b.strip()][:6]
    except Exception:
        selected = []
    result = None
    error = None
    try:
        da = _da()
        available = da.get_banks()
        selected = [b for b in selected if b in available]
        if selected:
            result = da.compare_banks(selected)
            if hasattr(result, "to_dict"):
                result = result.to_dict(orient="records")
    except Exception as e:
        error = str(e)
    return render_template("bank_compare.html", result=result, selected=selected,
                           presets=presets, error=error)


@routes_bp.route("/analytics/correlation")
@login_required
def correlation():
    matrix = None
    labels = None
    error = None
    try:
        df = _da().correlation_matrix()
        labels = [str(c) for c in df.columns]
        matrix = [[None if str(c) != str(r) else round(float(df.loc[r, c]), 2) for c in df.columns] for r in df.index]
    except Exception as e:
        error = str(e)
    return render_template("correlation.html", matrix=matrix, labels=labels, error=error)


@routes_bp.route("/analytics/personal")
@login_required
def personal_analytics():
    account = _acct()
    user = _user()
    c = _c()
    c.execute("SELECT COUNT(*) as n, COALESCE(SUM(amount),0) as tot, AVG(amount) as avg_amt FROM transactions WHERE user_id=? AND type IN ('withdraw','transfer')",
              (session["user_id"],))
    spend = dict(zip([d[0] for d in c.description], c.fetchone()))
    date_sql = "TO_CHAR(timestamp, 'YYYY-MM')" if _USE_PG else "strftime('%Y-%m', timestamp)"
    c.execute(f"""SELECT {date_sql} as month, type, SUM(amount) as total
                  FROM transactions WHERE user_id=? GROUP BY month, type ORDER BY month
                  LIMIT 6""", (session["user_id"],))
    cols = [d[0] for d in c.description]
    monthly = [dict(zip(cols, r)) for r in c.fetchall()]

    forecast = None
    error = None
    rep = _user_report() or {}
    forecast = rep.get("forecast")
    if not forecast:
        try:
            from banking_core.models import SpendingForecaster
            sf = SpendingForecaster()
            ok, res, err = _run_model(sf, lambda: sf.predict(session["user_id"], {
                "age": user["age"], "income_bracket": user.get("income_bracket"),
                "balance": account.balance}), sf.train)
            forecast = res if ok else None
            error = err if not ok else None
        except Exception as e:
            error = str(e)
    return render_template("personal_analytics.html", account=account, spend=spend,
                           monthly=monthly, forecast=forecast, error=error)


# ── One-Click Analysis Report ───────────────────────────────

@routes_bp.route("/analysis-report")
@login_required
def analysis_report():
    """Instant precomputed report: personal + industry, zero live model training."""
    account = _acct()
    rep = _user_report() or {}
    industry = _snapshot("analysis_report", kind="report") or {}
    if request.args.get("format") == "json":
        payload = {"personal": rep, "industry": industry}
        return Response(
            json.dumps(payload, default=str),
            mimetype="application/json",
            headers={"Content-Disposition": 'attachment; filename="analysis_report.json"'},
        )
    return render_template("analysis_report.html", account=account, rep=rep, industry=industry)


# ── ML Insights Hub ──────────────────────────────────────────

@routes_bp.route("/insights")
@login_required
def insights():
    account = _acct()
    if not account:
        return redirect(url_for("auth.logout"))
    feats = _feature_context()
    return render_template("insights.html", account=account, feats=feats)


@routes_bp.route("/ml/credit-prediction")
@login_required
def ml_credit_prediction():
    account = _acct()
    feats = _feature_context()
    prediction = {}
    rep = _user_report() or {}
    prediction = rep.get("credit_ml")
    if not prediction:
        try:
            from banking_core.models import CreditScorer
            cs = CreditScorer()
            ok, res, err = _run_model(cs, lambda: cs.predict({
                "age": feats["age"], "income_bracket": feats["income_bracket"],
                "balance": account.balance, "txn_count": feats["txn_count"]}), cs.train)
            prediction = res if ok else {"error": err, "fallback": account.credit_score}
        except Exception as e:
            prediction = {"error": str(e), "fallback": account.credit_score}
    return render_template("ml_credit.html", account=account, prediction=prediction, feats=feats)


@routes_bp.route("/ml/churn-analysis")
@login_required
def ml_churn_analysis():
    account = _acct()
    feats = _feature_context()
    churn = None
    rep = _user_report() or {}
    churn = rep.get("churn")
    if not churn:
        try:
            from banking_core.models import ChurnPredictor
            cp = ChurnPredictor()
            ok, res, err = _run_model(cp, lambda: cp.predict({
                "age": feats["age"], "income_bracket": feats["income_bracket"],
                "balance": account.balance, "txn_count": feats["txn_count"],
                "days_inactive": feats["days_inactive"]}), cp.train)
            churn = res if ok else {"error": err}
        except Exception as e:
            churn = {"error": str(e)}
    return render_template("ml_churn.html", account=account, churn=churn, feats=feats)


@routes_bp.route("/ml/loan-default", methods=["GET", "POST"])
@login_required
def ml_loan_default():
    account = _acct()
    feats = _feature_context()
    default_risk = None
    amortization = None
    loan_amount = float(request.form.get("amount", 200000)) if request.method == "POST" else 200000
    loan_rate = float(request.form.get("rate", 10.0)) if request.method == "POST" else 10.0
    tenure = int(request.form.get("tenure", 24)) if request.method == "POST" else 24
    if _USE_PG and request.method == "GET":
        rep = _user_report() or {}
        loan = rep.get("loan")
        if loan and loan.get("result"):
            default_risk = loan["result"]
            amortization = None
            return render_template("ml_loan_default.html", account=account, default_risk=default_risk,
                                   feats=feats, amortization=amortization, loan_amount=loan_amount,
                                   loan_rate=loan_rate, tenure=tenure)
    try:
        from banking_core.models import LoanDefaultModel
        lm = LoanDefaultModel()
        ok, res, err = _run_model(lm, lambda: lm.predict({
            "credit_score": feats["credit_score"], "balance": feats["balance"],
            "age": feats["age"], "income_bracket": feats["income_bracket"]},
            loan_amount, loan_rate, tenure), lm.train)
        default_risk = res if ok else {"error": err}
        if ok:
            amortization = lm.generate_amortization_schedule(loan_amount, loan_rate, tenure)
    except Exception as e:
        default_risk = {"error": str(e)}
    return render_template("ml_loan_default.html", account=account, default_risk=default_risk,
                           feats=feats, amortization=amortization, loan_amount=loan_amount,
                           loan_rate=loan_rate, tenure=tenure)


@routes_bp.route("/ml/bank-recommendation")
@login_required
def ml_bank_recommendation():
    account = _acct()
    feats = _feature_context()
    recommendation = {}
    rep = _user_report() or {}
    bankrec = rep.get("bankrec")
    if bankrec and bankrec.get("banks"):
        recommendation = {"banks": bankrec["banks"]}
    else:
        try:
            from banking_core.models import BankRecommender
            br = BankRecommender()
            ok, res, err = _run_model(br, lambda: br.recommend({
                "age": feats["age"], "income_bracket": feats["income_bracket"],
                "balance": account.balance, "bank": account.bank}, 5), br.train)
            recommendation = {"banks": res} if ok else {"error": err}
        except Exception as e:
            recommendation = {"error": str(e)}
    return render_template("ml_bank_rec.html", account=account, recommendation=recommendation, feats=feats)


# ── ML Lab (industry models) ─────────────────────────────────

def _bank_names_list():
    try:
        return _da().get_banks()
    except Exception:
        return ["STATE BANK OF INDIA", "HDFC BANK LTD"]


@routes_bp.route("/ml/cash-demand", methods=["GET", "POST"])
@login_required
def cash_demand():
    bank = request.args.get("bank") or (request.form.get("bank") if request.method == "POST" else "") or "STATE BANK OF INDIA"
    metric = request.args.get("metric") or (request.form.get("metric") if request.method == "POST" else "") or "DC_Vol_Cash_ATM"
    banks = _bank_names_list()
    if bank not in banks:
        bank = banks[0] if banks else bank
    prediction = None
    backtest = None
    error = None
    precomputed = False
    if _USE_PG:
        try:
            from banking_core.data.postgres_adapter import get_ml_snapshot
            snap = get_ml_snapshot("cash_demand", bank=bank, metric=metric, kind="forecast")
            if snap:
                prediction = snap
                precomputed = True
            else:
                error = f"No precomputed forecast for {bank} · {metric} on the deployed instance."
        except Exception as e:
            error = str(e)
        return render_template("ml_cash_demand.html", banks=banks, bank=bank, metric=metric,
                               prediction=prediction, backtest=None, error=error, precomputed=precomputed)
    try:
        from banking_core.models import CashDemandForecaster
        fc = CashDemandForecaster()
        ok, res, err = _run_model(fc, lambda: fc.predict(bank, metric),
                                  lambda: fc.train(bank, metric))
        prediction = res if ok else None
        error = err if not ok else None
        if ok:
            bt = fc.backtest(bank, metric)
            if isinstance(bt, dict) and "error" not in bt:
                backtest = bt
    except Exception as e:
        error = str(e)
    return render_template("ml_cash_demand.html", banks=banks, bank=bank, metric=metric,
                           prediction=prediction, backtest=backtest, error=error)


@routes_bp.route("/ml/txn-volume", methods=["GET", "POST"])
@login_required
def txn_volume():
    target = request.form.get("target") if request.method == "POST" else "DC_Vol_Cash_ATM"
    features = None
    importance = None
    error = None
    precomputed = False
    if _USE_PG:
        try:
            from banking_core.data.postgres_adapter import get_ml_snapshot
            snap = get_ml_snapshot("txn_volume", bank=None, metric=target, kind="predict")
            if snap:
                features = snap.get("features")
                importance = snap.get("importance")
                precomputed = True
            else:
                error = f"No precomputed prediction for target {target}."
        except Exception as e:
            error = str(e)
        return render_template("ml_txn_volume.html", features=features, importance=importance,
                               error=error, target=target, precomputed=precomputed)
    try:
        import pandas as pd
        from banking_core.models import TransactionPredictor
        tp = TransactionPredictor()
        demo = {"Total_Txn_Vol": 5_000_000, "Total_Cards": 3_000_000, "Digital_Share": 45,
                "Cash_Share": 55, "Total_ATMs": 1200, "PoS": 50000}

        def _predict():
            row = {c: 0.0 for c in tp.feature_cols}
            for k, v in demo.items():
                if k in row:
                    row[k] = v
            return tp.predict(pd.DataFrame([row]))

        ok, res, err = _run_model(tp, _predict, lambda: tp.train(target))
        features = res if ok else None
        error = err if not ok else None
        if ok:
            importance = tp.get_feature_importance()
    except Exception as e:
        error = str(e)
    return render_template("ml_txn_volume.html", features=features, importance=importance,
                           error=error, target=target)


@routes_bp.route("/ml/bank-clustering", methods=["GET", "POST"])
@login_required
def bank_clustering():
    k = request.form.get("k", 4, type=int) if request.method == "POST" else 4
    profiles = None
    optimal = None
    error = None
    if _USE_PG and k == 4:
        snap = _snapshot("analysis_clustering", kind="report")
        if snap:
            profiles = snap.get("profiles")
            optimal = snap.get("optimal_k")
            error = snap.get("error")
            return render_template("ml_clustering.html", profiles=profiles, optimal=optimal,
                                   error=error, k=k)
    try:
        from banking_core.models import BankClustering
        bc = BankClustering()
        try:
            optimal = bc.get_optimal_k(10)
        except Exception:
            optimal = None
        ok, res, err = _run_model(bc, lambda: bc.get_cluster_profiles(),
                                  lambda: bc.train(k=k))
        profiles = res if ok else None
        error = err if not ok else None
    except Exception as e:
        error = str(e)
    return render_template("ml_clustering.html", profiles=profiles, optimal=optimal,
                           error=error, k=k)


@routes_bp.route("/ml/anomaly-detection", methods=["GET", "POST"])
@login_required
def anomaly_detection():
    contamination = request.form.get("contamination", 0.05, type=float) if request.method == "POST" else 0.05
    flagged = None
    monthly = None
    error = None
    if _USE_PG and contamination == 0.05:
        snap = _snapshot("analysis_anomaly", kind="report")
        if snap:
            flagged = snap.get("flagged")
            monthly = snap.get("monthly")
            error = snap.get("error")
            return render_template("ml_anomaly.html", flagged=flagged, monthly=monthly,
                                   error=error, contamination=contamination)
    try:
        from banking_core.models import AnomalyDetector
        ad = AnomalyDetector()
        ok, res, err = _run_model(ad, lambda: ad.get_flagged_banks(15),
                                  lambda: ad.train(contamination=contamination))
        flagged = res if ok else None
        error = err if not ok else None
        if ok:
            monthly = ad.get_monthly_anomalies()
    except Exception as e:
        error = str(e)
    return render_template("ml_anomaly.html", flagged=flagged, monthly=monthly,
                           error=error, contamination=contamination)


@routes_bp.route("/ml/trend-decomposition", methods=["GET", "POST"])
@login_required
def trend_decomposition():
    bank = request.args.get("bank") or (request.form.get("bank") if request.method == "POST" else "") or "STATE BANK OF INDIA"
    metric = request.args.get("metric") or (request.form.get("metric") if request.method == "POST" else "") or "DC_Vol_Cash_ATM"
    banks = _bank_names_list()
    if bank not in banks:
        bank = banks[0] if banks else bank
    components = None
    error = None
    if _USE_PG:
        snap = _snapshot("analysis_trend", bank=bank, metric=metric, kind="decompose")
        if snap and snap.get("components"):
            return render_template("ml_decomposition.html", banks=banks, bank=bank, metric=metric,
                                   components=snap["components"], error=None)
        try:
            import pandas as pd
            from banking_core.data.postgres_adapter import get_industry_conn
            conn = get_industry_conn()
            df = pd.read_sql("SELECT * FROM atm_card_stats", conn)
            conn.close()
            data = df[df["Bank_Name"] == bank][["Reporting_Month", "Month_Num", metric]].sort_values("Month_Num")
            observed = data[metric].astype(float).tolist()
            months = data["Reporting_Month"].tolist()
            trend = data[metric].astype(float).rolling(3, min_periods=1).mean().tolist()
            components = {
                "months": [str(m) for m in months],
                "observed": [round(float(x), 2) if x is not None else None for x in observed],
                "trend": [None if pd.isna(x) else round(float(x), 2) for x in trend],
                "seasonal": [None] * len(months),
                "resid": [round(float(a) - (float(t) if not pd.isna(t) else 0), 2)
                          for a, t in zip(observed, trend)],
            }
            return render_template("ml_decomposition.html", banks=banks, bank=bank, metric=metric,
                                   components=components, error=None)
        except Exception as e:
            error = f"Precomputed decomposition unavailable: {e}"
            return render_template("ml_decomposition.html", banks=banks, bank=bank, metric=metric,
                                   components=None, error=error)
    try:
        from banking_core.models import TrendAnalyzer
        ta = TrendAnalyzer()
        ok, res, err = _run_model(ta, lambda: ta.decompose(bank, metric))
        error = err if not ok else None
        if ok:
            comps = res
            components = {
                "months": [d.strftime("%b") for d in comps.trend.index],
                "observed": [None if pd.isna(x) else round(float(x), 2) for x in comps.observed],
                "trend": [None if pd.isna(x) else round(float(x), 2) for x in comps.trend],
                "seasonal": [None if pd.isna(x) else round(float(x), 2) for x in comps.seasonal],
                "resid": [None if pd.isna(x) else round(float(x), 2) for x in comps.resid],
            }
    except Exception as e:
        error = str(e)
    return render_template("ml_decomposition.html", banks=banks, bank=bank, metric=metric,
                           components=components, error=error)


@routes_bp.route("/ml/channel-migration", methods=["GET", "POST"])
@login_required
def channel_migration():
    bank = request.args.get("bank") or (request.form.get("bank") if request.method == "POST" else "") or "STATE BANK OF INDIA"
    months = request.form.get("months", 6, type=int) if request.method == "POST" else 6
    banks = _bank_names_list()
    if bank not in banks:
        bank = banks[0] if banks else bank
    prediction = None
    error = None
    if _USE_PG and months == 6:
        snap = _snapshot("analysis_migration", bank=bank, kind="predict")
        if snap and snap.get("prediction"):
            prediction = snap["prediction"]
            return render_template("ml_migration.html", banks=banks, bank=bank, months=months,
                                   prediction=prediction, error=None)
    try:
        from banking_core.models import ChannelMigrationPredictor
        cm = ChannelMigrationPredictor()
        ok, res, err = _run_model(cm, lambda: cm.predict(bank, months),
                                  lambda: cm.train(bank))
        prediction = res if ok else None
        error = err if not ok else None
    except Exception as e:
        error = str(e)
    return render_template("ml_migration.html", banks=banks, bank=bank, months=months,
                           prediction=prediction, error=error)


@routes_bp.route("/ml/what-if", methods=["GET", "POST"])
@login_required
def what_if():
    changes = {}
    result = None
    error = None
    submitted = request.method == "POST"
    if _USE_PG and request.method == "GET":
        snap = _snapshot("analysis_whatif", kind="baseline")
        if snap:
            result = snap
            return render_template("ml_whatif.html", result=result, error=error, changes=changes,
                                   submitted=False)
    try:
        from banking_core.models import WhatIfSimulator
        ws = WhatIfSimulator()
        if submitted:
            for key in ("Total_ATMs", "PoS", "Digital_Share", "Total_Cards"):
                raw = request.form.get(key, "").strip()
                if raw:
                    try:
                        changes[key] = float(raw)
                    except ValueError:
                        pass
            if changes:
                ok, res, err = _run_model(ws, lambda: ws.simulate(changes), lambda: ws.train())
                result = res if ok else None
                error = err if not ok else None
    except Exception as e:
        error = str(e)
    return render_template("ml_whatif.html", result=result, error=error, changes=changes, submitted=submitted)


@routes_bp.route("/ml/replenishment", methods=["GET", "POST"])
@login_required
def replenishment():
    demand = request.form.get("demand", 100000, type=float) if request.method == "POST" else 100000
    capacity = request.form.get("capacity", 300000, type=float) if request.method == "POST" else 300000
    cost_per_visit = request.form.get("cost_per_visit", 500, type=float) if request.method == "POST" else 500
    holding_pct = request.form.get("holding_pct", 12, type=float) if request.method == "POST" else 12
    result = None
    comparison = None
    error = None
    try:
        from banking_core.models import ATMReplenishmentOptimizer
        opt = ATMReplenishmentOptimizer()
        result = opt.optimize(demand, capacity, cost_per_visit, holding_pct / 100)
        cmp = opt.compare_strategies(demand)
        comparison = cmp.get("strategies", [])
        best_strategy = cmp.get("best_strategy")
    except Exception as e:
        error = str(e)
    return render_template("ml_replenishment.html", result=result, comparison=comparison,
                           error=error, demand=demand, capacity=capacity,
                           cost_per_visit=cost_per_visit, holding_pct=holding_pct,
                           best_strategy=best_strategy)


@routes_bp.route("/ml/lstm-vs-prophet", methods=["GET", "POST"])
@login_required
def lstm_vs_prophet():
    bank = request.args.get("bank") or (request.form.get("bank") if request.method == "POST" else "") or "STATE BANK OF INDIA"
    metric = request.args.get("metric") or (request.form.get("metric") if request.method == "POST" else "") or "DC_Vol_Cash_ATM"
    banks = _bank_names_list()
    if bank not in banks:
        bank = banks[0] if banks else bank
    prophet_result = None
    lstm_result = None
    error = None
    precomputed = False
    if _USE_PG:
        try:
            from banking_core.data.postgres_adapter import get_ml_snapshot
            snap_p = get_ml_snapshot("cash_demand", bank=bank, metric=metric, kind="forecast")
            snap_l = get_ml_snapshot("lstm", bank=bank, metric=metric, kind="forecast")
            prophet_result = snap_p
            lstm_result = snap_l
            precomputed = bool(snap_p or snap_l)
            if not snap_p and not snap_l:
                error = f"No precomputed forecasts for {bank} · {metric} on the deployed instance."
        except Exception as e:
            error = str(e)
        return render_template("ml_lstm_vs_prophet.html", banks=banks, bank=bank, metric=metric,
                               prophet=prophet_result, lstm=lstm_result, error=error, precomputed=precomputed)
    try:
        from banking_core.models import CashDemandForecaster, LSTMForecaster
        fc = CashDemandForecaster()
        ok_p, res_p, err_p = _run_model(fc, lambda: fc.predict(bank, metric),
                                        lambda: fc.train(bank, metric))
        prophet_result = res_p if ok_p else None
        lstm = LSTMForecaster()
        ok_l, res_l, err_l = _run_model(lstm, lambda: lstm.predict(bank, metric),
                                        lambda: lstm.train(bank, metric))
        lstm_result = res_l if ok_l else None
        errors = [e for e in (err_p if not ok_p else None, err_l if not ok_l else None) if e]
        error = "; ".join(errors) if errors else None
    except Exception as e:
        error = str(e)
    return render_template("ml_lstm_vs_prophet.html", banks=banks, bank=bank, metric=metric,
                           prophet=prophet_result, lstm=lstm_result, error=error)


@routes_bp.route("/ml/retrain-all", methods=["GET", "POST"])
@login_required
def retrain_all():
    results = None
    if _USE_PG:
        results = [("Heavy models (Prophet / LSTM / XGBoost)", "Skipped",
                    "Precomputed snapshots are generated offline and stored in Neon")]
    elif request.method == "POST":
        from banking_core.models import (
            CashDemandForecaster, TransactionPredictor, BankClustering, AnomalyDetector,
            ChannelMigrationPredictor, WhatIfSimulator, CreditScorer, ChurnPredictor,
            LoanDefaultModel, SpendingForecaster, LSTMForecaster, TrendAnalyzer,
        )
        models = [
            ("Cash Demand (Prophet)", CashDemandForecaster(), lambda m: m.train()),
            ("LSTM Forecaster", LSTMForecaster(), lambda m: m.train()),
            ("Transaction Predictor", TransactionPredictor(), lambda m: m.train()),
            ("Bank Clustering", BankClustering(), lambda m: m.train()),
            ("Anomaly Detector", AnomalyDetector(), lambda m: m.train(contamination=0.05)),
            ("Channel Migration", ChannelMigrationPredictor(), lambda m: m.train()),
            ("What-If Simulator", WhatIfSimulator(), lambda m: m.train()),
            ("Trend Analyzer", TrendAnalyzer(), lambda m: {"skipped": "decomposition-only (statsmodels)"}),
            ("Credit Scorer", CreditScorer(), lambda m: m.train()),
            ("Churn Predictor", ChurnPredictor(), lambda m: m.train()),
            ("Loan Default", LoanDefaultModel(), lambda m: m.train()),
            ("Spending Forecaster", SpendingForecaster(), lambda m: m.train()),
        ]
        results = [_train_report(name, lambda m=m, fn=fn: fn(m)) for name, m, fn in models]
    return render_template("ml_retrain.html", results=results)


# ── Monitoring ───────────────────────────────────────────────

@routes_bp.route("/monitoring", methods=["GET", "POST"])
@login_required
def monitoring():
    versions = []
    stale_models = []
    train_outcome = None
    if request.method == "POST":
        model_name = request.form.get("model_name", "")
        import importlib
        mapping = {
            "cash_demand_forecaster": ("banking_core.models.cash_demand_forecaster", "CashDemandForecaster", ("STATE BANK OF INDIA", "DC_Vol_Cash_ATM")),
            "transaction_predictor": ("banking_core.models.transaction_predictor", "TransactionPredictor", ()),
            "bank_clustering": ("banking_core.models.bank_clustering", "BankClustering", ()),
            "anomaly_detector": ("banking_core.models.anomaly_detector", "AnomalyDetector", ()),
            "channel_migration": ("banking_core.models.channel_migration", "ChannelMigrationPredictor", ()),
            "what_if_simulator": ("banking_core.models.what_if_simulator", "WhatIfSimulator", ()),
            "credit_scorer": ("banking_core.models.credit_scorer", "CreditScorer", ()),
            "churn_predictor": ("banking_core.models.churn_predictor", "ChurnPredictor", ()),
            "loan_default_model": ("banking_core.models.loan_default_model", "LoanDefaultModel", ()),
            "spending_forecaster": ("banking_core.models.spending_forecaster", "SpendingForecaster", ()),
            "lstm_forecaster": ("banking_core.models.lstm_forecaster", "LSTMForecaster", ()),
        }
        spec = mapping.get(model_name)
        try:
            if _USE_PG:
                train_outcome = (model_name, "Skipped — precomputed on Neon")
            elif spec:
                mod = importlib.import_module(spec[0])
                cls = getattr(mod, spec[1])
                inst = cls()
                result = inst.train(*spec[2])
                if isinstance(result, dict) and result.get("error"):
                    train_outcome = (model_name, f"Failed: {result['error'][:80]}")
                else:
                    train_outcome = (model_name, "Trained")
            else:
                train_outcome = (model_name, "Unknown model")
        except Exception as e:
            train_outcome = (model_name, f"Failed: {str(e)[:80]}")
    try:
        from banking_core.models.model_monitor import ModelMonitor
        monitor = ModelMonitor()
        for model_name, info in monitor.get_all_model_status().items():
            versions.append({
                "model_name": model_name,
                "version": info.get("version", 0),
                "metrics": info.get("metrics_summary", {}),
                "last_trained": info.get("last_trained", ""),
                "freshness_days": info.get("freshness_days") or 999,
            })
            if info.get("freshness_days", 0) > 7:
                stale_models.append(model_name)
        versions.sort(key=lambda r: r["model_name"])
    except Exception:
        pass
    return render_template("monitoring.html", versions=versions, stale_models=stale_models,
                           train_outcome=train_outcome)


# ── Feedback ─────────────────────────────────────────────────

@routes_bp.route("/feedback", methods=["GET", "POST"])
@login_required
def feedback():
    analysis = None
    if request.method == "POST":
        rating = request.form.get("rating", "").strip()
        comments = request.form.get("comments", "").strip()
        category = request.form.get("category", "general").strip()
        if rating and rating.isdigit():
            _conn().execute("INSERT INTO feedback (user_id, rating, comments, category) VALUES (?,?,?,?)",
                            (session["user_id"], int(rating), comments, category))
            _conn().commit()
            try:
                from banking_core.models import SentimentAnalyzer
                sa = SentimentAnalyzer()
                analysis = sa.analyze_text(comments) if comments else None
            except Exception:
                analysis = None
            flash("Thank you for your feedback!" + (" Sentiment: " + analysis.get("sentiment_label") if analysis else ""), "success")
        else:
            flash("Rating is required", "error")
        return redirect(url_for("routes.feedback", analyzed=1))
    return render_template("feedback.html", analyzed=request.args.get("analyzed") == "1")


# ── Settings ─────────────────────────────────────────────────

@routes_bp.route("/settings")
@login_required
def settings_page():
    info = {}
    try:
        conn = _conn()
        c = conn.cursor()
        if _USE_PG:
            c.execute("SELECT table_name FROM information_schema.tables WHERE table_schema='public' ORDER BY table_name")
        else:
            c.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        tables = [r[0] for r in c.fetchall()]
        table_info = []
        for t in tables:
            try:
                c.execute(f"SELECT COUNT(*) FROM {t}")
                table_info.append({"name": t, "rows": c.fetchone()[0]})
            except Exception:
                table_info.append({"name": t, "rows": None})
        info["tables"] = table_info
    except Exception as e:
        info["error"] = str(e)

    data_status = {}
    try:
        from banking_core.utils import DATA_RAW, DATA_PROCESSED, DATA_MODELS, OUTPUTS_REPORTS, DATA_TRAINING
        for label, path in [("Raw data", DATA_RAW), ("Processed data", DATA_PROCESSED),
                            ("Models", DATA_MODELS), ("Reports", OUTPUTS_REPORTS), ("Training data", DATA_TRAINING)]:
            if path.exists():
                files = list(path.iterdir())
                data_status[label] = {"count": len(files), "size_mb": round(sum(f.stat().st_size for f in files if f.is_file()) / 1e6, 2)}
            else:
                data_status[label] = {"count": 0, "size_mb": 0}
    except Exception:
        pass

    rbi_status = None
    try:
        from banking_core.utils import DATA_RAW
        manifest = str(DATA_RAW / "sources_manifest.json")
        if os.path.exists(manifest):
            with open(manifest, encoding="utf-8") as f:
                rbi_status = json.load(f)
    except Exception:
        rbi_status = None
    return render_template("settings.html", info=info, data_status=data_status, rbi_status=rbi_status)


@routes_bp.route("/settings/refresh-rbi", methods=["POST"])
@login_required
def settings_refresh_rbi():
    try:
        from banking_core.utils import DATA_RAW, DATA_PROCESSED
        import subprocess, sys
        script = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
                              "v1_banking_core", "scripts", "download_rbi_data.py")
        if os.path.exists(script):
            env = dict(os.environ)
            env["PYTHONIOENCODING"] = "utf-8"
            proc = subprocess.run([sys.executable, script], capture_output=True, text=True,
                                  encoding="utf-8", timeout=300, env=env)
            tail = (proc.stdout or "")[-600:]
            flash(f"RBI data refresh done. {tail}", "success" if proc.returncode == 0 else "warning")
        else:
            flash("Download script not found", "error")
    except Exception as e:
        flash(f"Refresh failed: {e}", "error")
    return redirect(url_for("routes.settings_page"))


@routes_bp.route("/settings/reset-atm-usage", methods=["POST"])
@login_required
def settings_reset_atm():
    _usvc().reset_atm_usage()
    flash("ATM usage counters reset for all users", "success")
    return redirect(url_for("routes.settings_page"))


@routes_bp.route("/export/training-data")
@login_required
def export_training_data():
    try:
        from banking_core.utils import DATA_TRAINING
        from banking_core.data_generator import DataGenerator
        dg = DataGenerator()
        dg.export_all(scenario="WEB_EXPORT")
        dg.close()
        csvs = sorted([p for p in DATA_TRAINING.glob("*.csv")])
        if not csvs:
            flash("No training datasets generated", "warning")
            return redirect(url_for("routes.settings_page"))
        import zipfile
        bio = io.BytesIO()
        with zipfile.ZipFile(bio, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in csvs:
                zf.write(p, arcname=p.name)
        bio.seek(0)
        return Response(bio.getvalue(), mimetype="application/zip",
                        headers={"Content-Disposition": "attachment;filename=training_data.zip"})
    except Exception as e:
        flash(f"Export failed: {e}", "error")
    return redirect(url_for("routes.settings_page"))


# ── Demo Tour ────────────────────────────────────────────────

@routes_bp.route("/demo")
@login_required
def demo_tour():
    return render_template("demo_tour.html")


# ── Privacy & Data Rights ────────────────────────────────────

@routes_bp.route("/privacy")
def privacy_page():
    return render_template("privacy.html")


@routes_bp.route("/sessions")
@login_required
def sessions_page():
    user_id = session["user_id"]
    sessions_list = _usvc().get_sessions(user_id)
    current_token = session.get("session_token")
    for s in sessions_list:
        s["is_current"] = bool(s.get("token")) and s.get("token") == current_token
    return render_template("sessions.html", sessions=sessions_list)


@routes_bp.route("/sessions/<int:sid>/revoke", methods=["POST"])
@login_required
def session_revoke(sid):
    _usvc().revoke_session(session["user_id"], sid)
    flash("Session revoked", "success")
    return redirect(url_for("routes.sessions_page"))


@routes_bp.route("/sessions/revoke-all", methods=["POST"])
@login_required
def session_revoke_all():
    _usvc().revoke_all_sessions(session["user_id"])
    session.clear()
    flash("All sessions were ended. Please log in again.", "info")
    return redirect(url_for("auth.login"))


@routes_bp.route("/export/data")
@login_required
def export_my_data():
    bundle = _usvc().get_user_data_bundle(session["user_id"])
    body = json.dumps(bundle, indent=2, default=str)
    return Response(body, mimetype="application/json",
                    headers={"Content-Disposition":
                             "attachment;filename=my_data_export.json"})


@routes_bp.route("/delete-account", methods=["POST"])
@login_required
def delete_account():
    _usvc().delete_account(session["user_id"])
    session.clear()
    flash("Your account and all associated data have been permanently deleted.", "info")
    return redirect(url_for("auth.login"))
